"""Spreadsheet Q&A — read CSV / Excel and answer questions about the data via the LLM."""
from __future__ import annotations
import csv
import io
from pathlib import Path

from ..llm import make_llm_client

MAX_ROWS = 200          # cap rows fed to the LLM
MAX_CHARS = 20_000
_client = None


def _read_csv(path: Path) -> str:
    rows = []
    try:
        with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= MAX_ROWS:
                    rows.append(["...truncated..."])
                    break
                rows.append(row)
    except Exception as e:
        return f"(CSV read error: {e})"
    return "\n".join(", ".join(str(c) for c in r) for r in rows)


def _read_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "(openpyxl not installed — run: pip install openpyxl)"
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        out = []
        for ws in wb.worksheets:
            out.append(f"# Sheet: {ws.title}")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= MAX_ROWS:
                    out.append("...truncated...")
                    break
                out.append(", ".join("" if c is None else str(c) for c in row))
        wb.close()
        return "\n".join(out)
    except Exception as e:
        return f"(XLSX read error: {e})"


def analyze_spreadsheet(path: str, question: str = "") -> str:
    """Read a CSV/XLSX file and answer a question (or summarize the data)."""
    p = Path(path).expanduser()
    if not p.exists():
        return f"File not found: {p}"
    if not p.is_file():
        return f"Not a file: {p}"

    ext = p.suffix.lower()
    if ext == ".csv":
        data = _read_csv(p)
    elif ext in (".xlsx", ".xlsm"):
        data = _read_xlsx(p)
    else:
        return f"Unsupported spreadsheet type '{ext}'. Use .csv or .xlsx."

    if data.startswith("("):    # error sentinel
        return data
    if not data.strip():
        return f"{p.name} appears empty."
    if len(data) > MAX_CHARS:
        data = data[:MAX_CHARS] + "\n...(truncated)"

    global _client
    if _client is None:
        _client = make_llm_client()
    prompt = (
        f"Here is tabular data from a spreadsheet (comma-separated). "
        + ("Answer the question using ONLY this data. Show numbers precisely."
           if question else "Summarize what this data contains in 3-5 sentences.")
        + f"\n\n--- DATA ---\n{data}\n\n"
        + (f"--- QUESTION ---\n{question}\n" if question else "")
    )
    try:
        r = _client.chat(
            system="You are a precise data analyst. Compute carefully from the given table only.",
            history=[_client.make_user_message(prompt)],
            tools=[],
        )
        return (r.text or "").strip() or "(no answer)"
    except Exception as e:
        return f"Spreadsheet analysis failed: {e}"
