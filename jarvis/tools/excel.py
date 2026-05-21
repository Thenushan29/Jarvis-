"""Excel control — create, write, append, and read spreadsheet cells (openpyxl).

Reliable file-based Excel manipulation (no fragile COM). For driving the LIVE
Excel window, use the computer-use / automation tools instead.
"""
from __future__ import annotations
from pathlib import Path


def _wb(path: Path):
    from openpyxl import load_workbook, Workbook
    if path.exists():
        return load_workbook(str(path))
    return Workbook()


def create_excel(path: str, rows: list, sheet: str = "Sheet1") -> str:
    """Create a spreadsheet from rows (list of lists). Overwrites if exists."""
    from openpyxl import Workbook
    p = Path(path).expanduser()
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        p = p.with_suffix(".xlsx")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet
        for row in (rows or []):
            ws.append(list(row) if isinstance(row, (list, tuple)) else [row])
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(p))
        return f"Created {p} with {len(rows or [])} row(s)."
    except Exception as e:
        return f"Excel create failed: {e}"


def append_row(path: str, row: list, sheet: str = "") -> str:
    """Append one row to a spreadsheet (creates it if missing)."""
    p = Path(path).expanduser()
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        p = p.with_suffix(".xlsx")
    try:
        wb = _wb(p)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        ws.append(list(row) if isinstance(row, (list, tuple)) else [row])
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(p))
        return f"Appended a row to {p.name}."
    except Exception as e:
        return f"Append failed: {e}"


def set_cell(path: str, cell: str, value: str, sheet: str = "") -> str:
    """Set a single cell, e.g. cell='B2'."""
    p = Path(path).expanduser()
    try:
        wb = _wb(p)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        ws[cell] = value
        wb.save(str(p))
        return f"Set {cell} = {value} in {p.name}."
    except Exception as e:
        return f"Set cell failed: {e}"


def read_cell(path: str, cell: str, sheet: str = "") -> str:
    p = Path(path).expanduser()
    if not p.exists():
        return f"File not found: {p}"
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(p), data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        return f"{cell} = {ws[cell].value}"
    except Exception as e:
        return f"Read cell failed: {e}"
